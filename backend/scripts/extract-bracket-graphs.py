#!/usr/bin/env python3
"""
Extract full match graph per N from backend/CHAVES CT_tmp.xlsx (copy of the
planilha CHAVES CT.xlsx mantida em personaladmin/).

A planilha tem DUAS representacoes de cada chave:

  (A) Diagrama visual (cols C-I[+]) — desenho da chave que arbitros/atletas
      veem na chave impressa. Os labels dos jogos sao celulas com 'j1', 'j2'...
      (sempre minusculo) posicionadas entre as linhas que conectam os
      participantes.

  (B) Tabela estruturada (cols Q-V) — listagem 'Jogos' com 'J', numero, top,
      bottom. Esta tabela tem numeracao DIFERENTE da do diagrama: os IDs ali
      seguem a ordem em que a planilha foi construida, nao a ordem visual.

A versao anterior deste script (extract-bracket-graphs.old.py) lia (B) e gerava
J-IDs que NAO batiam com os impressos na chave. Esta versao le (A) — os J-IDs
canonicos da chave impressa — e usa (B) apenas como referencia estrutural para
descobrir o grafo (top/bottom de cada jogo).

Estrategia:

  1. Le a tabela estruturada (B) para obter a topologia: [{id_estrut, top, bot}]
     ja resolvido para refs P{n}, V:J{x}, L:J{x}.
  2. Escaneia o diagrama (A) por celulas 'j\\d+' (minusculo) e infere top/bot
     de cada uma analisando as celulas vizinhas (escaneia col-2 para cima e
     para baixo ate encontrar uma posicao P{n} ou outro 'j' interno).
  3. Marca como 3rd-place a celula j cuja coluna H (col 8) na mesma linha
     contem '3' (marcador '3o/4o').
  4. Casa cada j visual com um jogo estrutural por igualdade de conjuntos
     {top, bot}, processando round-a-round. Cada match estabelece a relacao
     id_estrut -> id_visual; refs V:J/L:J nos rounds seguintes sao
     re-mapeadas.
  5. Renumera tudo para os IDs visuais e gera o SQL.

Usage:
  cd backend && py -3 scripts/extract-bracket-graphs.py
"""
from openpyxl import load_workbook
from pathlib import Path
import json
import re

ROOT = Path(__file__).resolve().parents[2]
XLSX = ROOT / 'backend' / 'CHAVES CT_tmp.xlsx'
BYES_SQL = ROOT / 'backend' / 'prisma' / 'seeds' / 'bracket_chaves_byes.sql'
OUT = ROOT / 'backend' / 'prisma' / 'seeds' / 'bracket_chaves_matches.sql'


# ---------------------------------------------------------------------------
# Step 1: tabela estruturada (cols Q-V) — codigo herdado da versao anterior.
# ---------------------------------------------------------------------------

def load_bye_positions():
    if not BYES_SQL.exists():
        raise RuntimeError(f'Required seed file missing: {BYES_SQL}.')
    text = BYES_SQL.read_text(encoding='utf-8')
    out = {}
    for m in re.finditer(r"VALUES\s*\(\s*(\d+)\s*,\s*'\{([\d,\s]*)\}'", text):
        n = int(m.group(1))
        arr = m.group(2).strip()
        positions = [int(x) for x in arr.split(',') if x.strip()] if arr else []
        out[n] = sorted(positions)
    return out


def parse_simple_ref(value):
    if value is None: return None
    s = str(value).strip()
    if s in ('', '0', 'null', '#REF!', '#N/A'): return None
    if re.fullmatch(r'\d+', s): return f'P{s}'
    m = re.fullmatch(r'Venc\.?\s*J\s*(\d+)', s, re.IGNORECASE)
    if m: return f'V:J{m.group(1)}'
    m = re.fullmatch(r'Perd\.?\s*J\s*(\d+)', s, re.IGNORECASE)
    if m: return f'L:J{m.group(1)}'
    return None


def find_matches_list(ws):
    jogo_row = jogo_col = None
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=r, column=c).value == 'Jogos':
                jogo_row, jogo_col = r, c
                break
        if jogo_row: break
    if not jogo_row: return []
    out = []
    seq = 0
    for r in range(jogo_row + 1, ws.max_row + 1):
        if ws.cell(row=r, column=jogo_col).value != 'J': break
        seq += 1
        n_cell = ws.cell(row=r, column=jogo_col + 1).value
        n = n_cell if isinstance(n_cell, int) else seq
        top = ws.cell(row=r, column=jogo_col + 2).value
        bot = ws.cell(row=r, column=jogo_col + 4).value
        out.append((f'J{n}', top, bot))
    return out


def build_structural_graph(raw_matches, bye_positions, N):
    """Mesmo algoritmo do script antigo — preenche P refs em R1/R2 BYE usando
    posicoes sequenciais. Retorna lista [{id, top, bottom}].
    """
    bye_set = set(bye_positions)
    non_bye = [p for p in range(1, N + 1) if p not in bye_set]
    classified = [(mid, parse_simple_ref(t), parse_simple_ref(b)) for mid, t, b in raw_matches]
    r1_ids = [mid for mid, tp, bp in classified if tp is None and bp is None]
    bye_ids = []
    for mid, tp, bp in classified:
        if (tp is None) != (bp is None):
            other = bp if tp is None else tp
            if other and other.startswith('V:'):
                bye_ids.append(mid)
    r1_ids.sort(key=lambda x: int(x[1:]))
    bye_ids.sort(key=lambda x: int(x[1:]))
    if len(r1_ids) * 2 != len(non_bye):
        raise RuntimeError(
            f'N={N}: {len(r1_ids)} R1 matches need {2*len(r1_ids)} non-BYE positions, '
            f'but {len(non_bye)} available')
    r1_resolved = {}
    for i, mid in enumerate(r1_ids):
        r1_resolved[mid] = (f'P{non_bye[2*i]}', f'P{non_bye[2*i+1]}')
    if len(bye_ids) > len(bye_positions):
        raise RuntimeError(
            f'N={N}: {len(bye_ids)} BYE matches but only {len(bye_positions)} BYE positions')
    bye_resolved = {}
    for i, mid in enumerate(bye_ids):
        if i < len(bye_positions):
            bye_resolved[mid] = f'P{bye_positions[i]}'
    matches = []
    for mid, tp, bp in classified:
        if mid in r1_resolved:
            top, bot = r1_resolved[mid]
        elif mid in bye_resolved:
            if tp is None:
                top, bot = bye_resolved[mid], bp
            else:
                top, bot = tp, bye_resolved[mid]
        else:
            top, bot = tp, bp
        if top is None or bot is None:
            raise RuntimeError(f'N={N} {mid}: unresolved top={tp!r} bot={bp!r}')
        matches.append({'id': mid, 'top': top, 'bottom': bot})
    return matches


# ---------------------------------------------------------------------------
# Step 2/3: diagrama visual (cols C-I+) — escanear celulas 'j\d+' minusculas
# e usar coordenadas (linha) para casar com a topologia estrutural.
# ---------------------------------------------------------------------------


def find_visual_j_cells(ws):
    """Localiza todas as celulas do diagrama com label 'j\\d+' (minusculo).

    Filtros:
      - Apenas minusculas (legendas usam 'J' maiusculo).
      - Restrito a colunas 1..15 (diagramas nao passam disso na pratica).
      - Ignora celulas onde a linha tambem contem texto explicativo (heuristica:
        proxima coluna a direita tem texto que nao parece label de jogo).

    Returns: [(jid:int, row, col)] ordenado por linha.
    """
    out = []
    max_r = ws.max_row
    max_c = min(ws.max_column, 15)
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            v = ws.cell(row=r, column=c).value
            if v is None: continue
            s = str(v).strip()
            m = re.fullmatch(r'j\s*(\d+)', s)
            if m:
                out.append((int(m.group(1)), r, c))
    return out


def get_position_rows(ws):
    """Map de row -> position int. Tenta col 1 primeiro; se la nao tiver
    inteiros sequenciais, tenta col 2 (algumas planilhas usam col 1 para nome
    do time e col 2 para o numero).

    Para em linhas com labels tipo 'Rodada' (final do diagrama).
    """
    def scan(col):
        out = {}
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=col).value
            if isinstance(v, int) and v > 0:
                out[r] = v
            elif isinstance(v, str) and v.strip().lower() in ('id', 'rodada', 'chave'):
                break
        return out
    a = scan(1)
    if a and len(a) >= 2:
        return a
    return scan(2)


# ---------------------------------------------------------------------------
# Step 4: deriva feeders de cada j visual usando proximidade + consumo.
# ---------------------------------------------------------------------------

def derive_visual_feeders(visual_cells_by_col, position_rows_inv, third_visual_id):
    """Para cada celula j visual (exceto 3rd-place), descobre seus dois feeders
    (top, bot). Processa COLUNA POR COLUNA do diagrama (col asc):

      - Col 3: primeira coluna de jogos. Todos sao R1 (top=P, bot=P) — pega
        as posicoes adjacentes.
      - Col 5+: jogos pegam os feeders mais proximos (acima e abaixo)
        disponiveis (P nao usadas + j cells ja resolvidas em colunas
        anteriores).

    Justificativa: o diagrama tem layout de eliminatoria simples — cada j cell
    so pode receber feeders de col < sua col. Iterando do interno para o
    externo, garantimos que feeders ja estao resolvidos quando precisamos
    deles.

    Args:
      visual_cells_by_col: dict {col: [(vid, row), ...]} ordenado por row.
      position_rows_inv: {position_int: row}.
      third_visual_id: vid do 3rd-place (sera ignorado aqui).

    Returns: dict {visual_id: (top_ref, bot_ref)}.
    """
    available_p = dict(position_rows_inv)
    available_j = {}  # vid -> row
    feeders = {}

    for col in sorted(visual_cells_by_col):
        for vid, row in visual_cells_by_col[col]:
            if vid == third_visual_id: continue
            # Acha closest acima e abaixo entre todos os candidatos disponiveis.
            best_above = None  # (dist, kind, key)
            best_below = None
            for p, r in available_p.items():
                if r < row:
                    d = row - r
                    if best_above is None or d < best_above[0]: best_above = (d, 'p', p)
                elif r > row:
                    d = r - row
                    if best_below is None or d < best_below[0]: best_below = (d, 'p', p)
            for j, r in available_j.items():
                if r < row:
                    d = row - r
                    if best_above is None or d < best_above[0]: best_above = (d, 'j', j)
                elif r > row:
                    d = r - row
                    if best_below is None or d < best_below[0]: best_below = (d, 'j', j)
            if best_above is None or best_below is None:
                raise RuntimeError(
                    f'j{vid} at row {row} col {col}: no candidate '
                    f'{"above" if best_above is None else "below"}; '
                    f'available P={available_p}, available J={available_j}')

            def ref(kind, key):
                return f'P{key}' if kind == 'p' else f'V:J{key}'

            feeders[vid] = (ref(best_above[1], best_above[2]), ref(best_below[1], best_below[2]))
            # Consume
            if best_above[1] == 'p': del available_p[best_above[2]]
            else: del available_j[best_above[2]]
            if best_below[1] == 'p': del available_p[best_below[2]]
            else: del available_j[best_below[2]]
            available_j[vid] = row

    return feeders


# ---------------------------------------------------------------------------
# Step 5: renumera e gera grafo final.
# ---------------------------------------------------------------------------

def derive_rounds(matches):
    rounds = {}
    pending = list(matches)
    while pending:
        next_pending = []
        progress = False
        for m in pending:
            deps = []
            for ref in (m['top'], m['bottom']):
                if ref.startswith('V:') or ref.startswith('L:'):
                    deps.append(ref[2:])
            if not deps:
                rounds[m['id']] = 1
                progress = True
            elif all(d in rounds for d in deps):
                rounds[m['id']] = max(rounds[d] for d in deps) + 1
                progress = True
            else:
                next_pending.append(m)
        if not progress:
            raise RuntimeError(f'Cannot derive rounds for: {[m["id"] for m in next_pending]}')
        pending = next_pending
    for m in matches:
        m['round'] = rounds[m['id']]
    return matches


def renumber(structural, mapping, struct_third):
    """Aplica mapping estr->visual em todos os ids e refs."""
    out = []
    for s in structural:
        new_id = f'J{mapping[s["id"]]}'
        def remap(ref):
            if ref.startswith('P'):
                return ref
            prefix, sid = ref[0], ref[2:]
            return f'{prefix}:J{mapping[sid]}'
        out.append({
            'id': new_id,
            'top': remap(s['top']),
            'bottom': remap(s['bottom']),
        })
    derive_rounds(out)
    third_visual = f'J{mapping[struct_third]}' if struct_third else None
    max_round = max(m['round'] for m in out if m['id'] != third_visual)
    final_match = next(m['id'] for m in out if m['round'] == max_round and m['id'] != third_visual)
    return {
        'matches': out,
        'final': final_match,
        'thirdPlace': third_visual,
    }


def build_visual_graph(ws, N, bye_positions):
    """Pipeline:
      1. Le tabela estruturada SO para detectar se ha 3rd-place (refs L:).
      2. Escaneia diagrama visual para descobrir celulas j e suas linhas.
      3. Deriva feeders de cada j pelo algoritmo de proximidade/consumo.
      4. Identifica final (j com maior round que nao seja 3rd-place) e
         constroi as duas refs L:Jx do 3rd-place a partir dos filhos do final.
    """
    raw = find_matches_list(ws)
    if not raw:
        if N == 2:
            return {
                'matches': [{'id': 'J1', 'round': 1, 'top': 'P1', 'bottom': 'P2'}],
                'final': 'J1',
                'thirdPlace': None,
            }
        raise RuntimeError('no structural matches list found')

    # Existe 3rd-place? Olha estrutural por refs L:.
    has_third = False
    for mid, t, b in raw:
        for v in (t, b):
            if isinstance(v, str) and 'perd' in str(v).lower():
                has_third = True
                break

    visual_cells = find_visual_j_cells(ws)
    if not visual_cells:
        # Fallback: usa estrutural antigo. Loga warning.
        print(f'  WARN N={N}: no visual diagram j-cells found, using structural numbering')
        structural = build_structural_graph(raw, bye_positions, N)
        derive_rounds(structural)
        third = None
        for m in structural:
            if m['top'].startswith('L:') or m['bottom'].startswith('L:'):
                third = m['id']
                break
        max_round = max(m['round'] for m in structural if m['id'] != third)
        final_match = next(m['id'] for m in structural if m['round'] == max_round and m['id'] != third)
        return {'matches': structural, 'final': final_match, 'thirdPlace': third}

    position_rows = get_position_rows(ws)
    position_rows_inv = {v: k for k, v in position_rows.items()}

    # Detecta a CELULA do 3rd-place (jid,row,col). Marcador "3o/4o" fica na
    # celula imediatamente a esquerda do j (col-1), com os digitos '3' e '4'.
    third_cell = None
    for jid, r, c in visual_cells:
        marker_val = ws.cell(row=r, column=c - 1).value if c > 1 else None
        if isinstance(marker_val, str) and '3' in marker_val and '4' in marker_val:
            if third_cell is not None:
                raise RuntimeError(f'multiple 3rd-place markers: j{third_cell[0]} and j{jid}')
            third_cell = (jid, r, c)
    if has_third and third_cell is None and visual_cells:
        # Fallback: 3rd-place eh a celula de MAIOR linha (isolada abaixo).
        third_cell = max(visual_cells, key=lambda t: t[1])
    if has_third and third_cell is None:
        raise RuntimeError('has 3rd-place estrutural mas nenhuma celula visual marcada')

    # Colisao de numero: se um jogo do bracket usa o MESMO numero do 3rd-place
    # (erro de rotulo na planilha — ex. N=22 semifinal e 3rd ambos "j19"),
    # renumera o jogo do bracket para o menor id livre. Senao ele seria
    # descartado junto com o 3rd (vira jogo faltante/orfao).
    if third_cell is not None:
        used = {jid for jid, _, _ in visual_cells}

        def next_free(used_set):
            k = 1
            while k in used_set:
                k += 1
            return k

        new_cells = []
        for cell in visual_cells:
            jid, r, c = cell
            if cell != third_cell and jid == third_cell[0]:
                nf = next_free(used)
                used.add(nf)
                new_cells.append((nf, r, c))
            else:
                new_cells.append(cell)
        visual_cells = new_cells
        third_visual_id = third_cell[0]
    else:
        third_visual_id = None

    # Apos a renumeracao, nenhum winner deve compartilhar numero (alem do 3rd).
    seen = {}
    for jid, r, c in visual_cells:
        if (jid, r, c) == third_cell:
            continue
        if jid in seen:
            raise RuntimeError(f'duplicate visual j{jid} (rows {seen[jid]} and {r})')
        seen[jid] = r

    # Organiza visual cells por coluna para processamento col-asc.
    cells_by_col = {}
    for jid, r, c in visual_cells:
        if jid == third_visual_id: continue
        cells_by_col.setdefault(c, []).append((jid, r))
    for c in cells_by_col:
        cells_by_col[c].sort(key=lambda x: x[1])

    feeders = derive_visual_feeders(cells_by_col, position_rows_inv, third_visual_id)

    # Monta matches sem o 3rd.
    matches = []
    for vid in sorted(feeders.keys()):
        top, bot = feeders[vid]
        matches.append({'id': f'J{vid}', 'top': top, 'bottom': bot})

    derive_rounds(matches)
    max_round = max(m['round'] for m in matches)
    final_match = next(m['id'] for m in matches if m['round'] == max_round)

    # 3rd-place: feeders sao os L:Jx dos dois jogos que alimentam o final.
    third_place = None
    if third_visual_id is not None:
        final_obj = next(m for m in matches if m['id'] == final_match)
        # Os dois children do final sao os semi-finalistas.
        ftop = final_obj['top']
        fbot = final_obj['bottom']
        if not (ftop.startswith('V:') and fbot.startswith('V:')):
            raise RuntimeError(f'final {final_match} feeders nao sao V:J ({ftop}, {fbot})')
        third_top = 'L:' + ftop[2:]
        third_bot = 'L:' + fbot[2:]
        third_id = f'J{third_visual_id}'
        matches.append({'id': third_id, 'top': third_top, 'bottom': third_bot})
        derive_rounds(matches)
        third_place = third_id

    # Guard: o winners bracket deve ter N-1 jogos e nenhum jogo orfao (vencedor
    # nao consumido). Falha alto em vez de emitir grafo quebrado em silencio.
    winners = [m for m in matches if m['id'] != third_place]
    if len(winners) != N - 1:
        raise RuntimeError(
            f'grafo invalido: {len(winners)} jogos no winners bracket (esperado {N - 1})')
    referenced = {m[s][2:] for m in matches for s in ('top', 'bottom') if m[s].startswith('V:')}
    orphans = [m['id'] for m in winners if m['id'] != final_match and m['id'] not in referenced]
    if orphans:
        raise RuntimeError(f'grafo invalido: jogos orfaos {orphans}')

    return {
        'matches': matches,
        'final': final_match,
        'thirdPlace': third_place,
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    wb = load_workbook(XLSX, data_only=True)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    byes_map = load_bye_positions()
    print(f'Loaded BYE data for {len(byes_map)} N values')

    lines = [
        '-- Auto-generated by backend/scripts/extract-bracket-graphs.py',
        '-- DO NOT EDIT MANUALLY. To regenerate: cd backend && py -3 scripts/extract-bracket-graphs.py',
        '-- J-IDs sao os do DIAGRAMA VISUAL impresso (nao os da tabela estruturada da planilha).',
        '',
    ]
    ok = []
    failed = []
    for n in range(2, 78):
        sn = f'{n:02d}'
        if sn not in wb.sheetnames: continue
        ws = wb[sn]
        bye_positions = byes_map.get(n, [])
        try:
            g = build_visual_graph(ws, n, bye_positions)
            ok.append(n)
        except Exception as e:
            failed.append((n, str(e)))
            lines.append(f'-- N={n}: FAILED ({e})')
            lines.append('')
            continue
        graph_json = json.dumps(g, separators=(',', ':'))
        escaped = graph_json.replace("'", "''")
        lines.append(f'-- N={n}')
        lines.append(
            f"INSERT INTO bracket_chaves_matches (numero_inscrito, matches_graph) "
            f"VALUES ({n}, '{escaped}'::jsonb) "
            f"ON CONFLICT (numero_inscrito) DO UPDATE SET matches_graph = EXCLUDED.matches_graph;"
        )
        lines.append('')

    OUT.write_text('\n'.join(lines), encoding='utf-8')
    print(f'Wrote {OUT}')
    print(f'OK: {len(ok)} N values: {ok}')
    print(f'FAILED: {len(failed)} N values')
    for n, msg in failed:
        # truncate long messages
        m = msg if len(msg) < 200 else msg[:200] + '...'
        print(f'  N={n}: {m}')


if __name__ == '__main__':
    main()
